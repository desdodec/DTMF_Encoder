from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "names_data"
OUTPUT_PATH = ROOT / "names.json"
NAME_RE = re.compile(r"^[A-Za-z]+$")


def normalize_name(value):
    if value is None:
        return None

    name = str(value).strip().upper()
    if not NAME_RE.fullmatch(name):
        return None
    return name


def numeric_count(value):
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return 0


def add_record(names, raw_name, count, source, sex):
    name = normalize_name(raw_name)
    if not name:
        return False

    entry = names[name]
    entry["score"] += max(numeric_count(count), 1)
    entry["sources"].add(source)
    if sex:
        entry["sexes"].add(sex)
    return True


def locate_header(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        lowered = [str(value).strip().lower() if value is not None else "" for value in row]
        if "name" in lowered:
            return row_index, lowered
    return None, []


def process_historical_ons(path, names):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    processed = 0
    skipped = 0

    for sheet_name in ("Table_1", "Table_2"):
        ws = workbook[sheet_name]
        header_row, headers = locate_header(ws)
        if not header_row:
            continue

        name_index = headers.index("name")
        count_indexes = [index for index, header in enumerate(headers) if header.endswith(" count")]
        sex = "F" if sheet_name == "Table_1" else "M"

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            raw_name = row[name_index] if name_index < len(row) else None
            total = sum(numeric_count(row[index]) for index in count_indexes if index < len(row))
            if add_record(names, raw_name, total, path.name, sex):
                processed += 1
            else:
                skipped += 1

    return processed, skipped


def process_ons_2024(path, names, sex):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    processed = 0
    skipped = 0

    # Table_6 is the all-names England and Wales table, avoiding regional/month duplicates.
    ws = workbook["Table_6"]
    header_row, headers = locate_header(ws)
    if not header_row:
        return processed, skipped

    name_index = headers.index("name")
    count_index = headers.index("count") if "count" in headers else None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_name = row[name_index] if name_index < len(row) else None
        count = row[count_index] if count_index is not None and count_index < len(row) else 0
        if add_record(names, raw_name, count, path.name, sex):
            processed += 1
        else:
            skipped += 1

    return processed, skipped


def process_scotland_csv(path, names):
    processed = 0
    skipped = 0

    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            sex = {"BOY": "M", "GIRL": "F"}.get(row.get("Sex", "").strip().upper())
            if add_record(names, row.get("Name"), row.get("Number"), path.name, sex):
                processed += 1
            else:
                skipped += 1

    return processed, skipped


def main():
    names = defaultdict(lambda: {"score": 0, "sources": set(), "sexes": set()})
    source_stats = []

    processors = {
        "en_wales_babynames1996to2024.xlsx": lambda path: process_historical_ons(path, names),
        "en_wales_boysnames2024.xlsx": lambda path: process_ons_2024(path, names, "M"),
        "en_wales_girlsnames2024.xlsx": lambda path: process_ons_2024(path, names, "F"),
        "scotland_names_full-list-1974-2024.csv": lambda path: process_scotland_csv(path, names),
    }

    for filename, processor in processors.items():
        path = SOURCE_DIR / filename
        if not path.exists():
            source_stats.append(
                {
                    "file": filename,
                    "status": "skipped",
                    "processedRows": 0,
                    "skippedRows": 0,
                    "reason": "Source file was not found.",
                }
            )
            continue

        processed, skipped = processor(path)
        source_stats.append(
            {
                "file": filename,
                "status": "processed",
                "processedRows": processed,
                "skippedRows": skipped,
                "reason": None,
            }
        )

    ranked_names = [
        {
            "name": name,
            "score": data["score"],
            "sexes": sorted(data["sexes"]),
            "sources": sorted(data["sources"]),
        }
        for name, data in names.items()
    ]
    ranked_names.sort(key=lambda item: (-item["score"], item["name"]))

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "normalization": "Uppercase A-Z first names only; names containing spaces, hyphens, apostrophes, dots, accents, or other symbols are excluded to match app validation.",
        "sourceStats": source_stats,
        "totalNames": len(ranked_names),
        "names": ranked_names,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(json.dumps({"output": str(OUTPUT_PATH), "totalNames": len(ranked_names), "sourceStats": source_stats}, indent=2))


if __name__ == "__main__":
    main()
